"""
Sprint 2 — expense and revenue data cleaning.
Sprint 3 addendum — vendor-based category-consistency fixes and the
sparse-category handling rule for modeling.

Consolidates the 12 real monthly Master Sheets (Aug 2025-Jul 2026) into two
clean, private CSVs: cleaned_expenses.csv and cleaned_revenue.csv.

Every rule implemented here is documented in docs/data_dictionary.md under
"Resolved Cleaning Rules" - this file is the code version of that document,
and the two should stay in sync if a rule ever changes.

PRIVACY NOTE: this script contains no real vendor names, amounts, or
person names. The two spots that need real-data-specific values
(estimated dates for a handful of blank-date rows, and vendor-based
category corrections) import them from config/private_cleaning_overrides.py,
a git-ignored file - see config/private_cleaning_overrides.example.py for
the format. Without that file present, both are empty and this script
simply runs without those specific fixes, which is exactly the right
behavior for the public repo / synthetic demo path, where no real vendor
data exists to begin with.

Usage:
    python src/data_cleaning.py --raw-dir /path/to/extracted/exports

`--raw-dir` should point at a folder that directly contains the "2025" and
"2026" subfolders exactly as they exist in your original export zips. The
script does not modify, move, or rename anything under that folder - it only
reads from it.

Output goes to data/private/ (already git-ignored):
    data/private/cleaned_expenses.csv
    data/private/cleaned_revenue.csv
"""

import argparse
import datetime as dt
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

# Real-data-specific overrides (vendor names, amounts, person names) live
# in a git-ignored config file, never in this script. See the module
# docstring above and config/private_cleaning_overrides.example.py.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    from config.private_cleaning_overrides import ESTIMATED_DATES, VENDOR_CATEGORY_OVERRIDES
except ImportError:
    ESTIMATED_DATES = {}
    VENDOR_CATEGORY_OVERRIDES = []

# ---------------------------------------------------------------------------
# Category crosswalk: legacy Aug-Dec 2025 labels -> final 13-category scheme.
# See docs/data_dictionary.md "Legacy Category Crosswalk" for the reasoning
# behind each mapping. These are business category names, not vendor/person
# names, so they're safe to keep in the public script.
# ---------------------------------------------------------------------------

FINAL_CATEGORIES = [
    "Payroll & Labor",
    "Occupancy Cost",
    "Marketing & Growth",
    "Software & Technology",
    "Insurance",
    "Professional Services",
    "Licensing & Compliance",
    "Banking & Financial Fees",
    "Repairs, Maintenance & Small Equipment",
    "Debt Services",
    "Cost of Services (Clinical Supplies)",
    "Owner Distribution (Non-Operating)",
    "Other/Non-Operating",
]

CATEGORY_CROSSWALK = {
    # Already-final categories (Jan 2026 onward) map to themselves.
    "Payroll & Labor": "Payroll & Labor",
    "Occupancy Cost": "Occupancy Cost",
    "Marketing & Growth": "Marketing & Growth",
    "Software & Technology": "Software & Technology",
    "Insurance": "Insurance",
    "Professional Services": "Professional Services",
    "Licensing & Compliance": "Licensing & Compliance",
    "Banking & Financial Fees": "Banking & Financial Fees",
    "Repairs, Maintence & Small Equipment": "Repairs, Maintenance & Small Equipment",
    "Debt Services": "Debt Services",
    "Cost of Services (Clinical Supplies)": "Cost of Services (Clinical Supplies)",
    "Owner Distrubtion (Non-Operating)": "Owner Distribution (Non-Operating)",
    "Other/ Non-Operating": "Other/Non-Operating",
    "Donation": "Other/Non-Operating",
    # Legacy Aug-Dec 2025 granular scheme.
    "Salaries & Wages": "Payroll & Labor",
    "Salaries & Wages ": "Payroll & Labor",  # trailing-space variant seen in source
    "Payroll Taxes": "Payroll & Labor",
    "Staff Meals": "Payroll & Labor",
    "Employee Reimbursement - Staff Meals": "Payroll & Labor",
    "Contract Labor": "Payroll & Labor",
    "Contacted Work": "Payroll & Labor",
    "Practice-Mgmt Software": "Software & Technology",
    "General SaaS": "Software & Technology",
    "Website & SEO": "Marketing & Growth",
    "Digital Ads": "Marketing & Growth",
    "Print/ In house Promo": "Marketing & Growth",
    "Rent/ CAM": "Occupancy Cost",
    "Utilities": "Occupancy Cost",
    "Disposable Clinic Supplies": "Cost of Services (Clinical Supplies)",
    "Skincare & Peel Products": "Cost of Services (Clinical Supplies)",
    "Injectable & IV Supplies": "Cost of Services (Clinical Supplies)",
    "Bank & Merchant Fees": "Banking & Financial Fees",
    "Interest Expense": "Debt Services",
    "Owner Draw/ Distribution": "Owner Distribution (Non-Operating)",
    "Insurance-Malpratice & Liability": "Insurance",
    "Licenses & Permits": "Licensing & Compliance",
    "Repairs & Maintenance": "Repairs, Maintenance & Small Equipment",
    "Small Tools & Décor": "Repairs, Maintenance & Small Equipment",
    "Employee Reimbursement - Office Supplies": "Other/Non-Operating",
    "Office Supplies": "Other/Non-Operating",
    "Machine Payment": "Other/Non-Operating",
    "Charitable Contributions": "Other/Non-Operating",
    "Deprieciation & Amortization": "Other/Non-Operating",
    "misc": "Other/Non-Operating",
}

# ---------------------------------------------------------------------------
# Source file manifest - see docs/data_dictionary.md "Source File Manifest"
# for why each of these was picked over its duplicate/partial candidates.
# These are relative file paths within your own raw export folder, not
# published anywhere else, and contain no content themselves.
# ---------------------------------------------------------------------------

EXPENSE_SOURCES = {
    "2025-08": ("2025/Taxes 2025/Cosmedici Laser/Accounted Statements/August Master Sheet.xlsx", "August_Master_Sheet (version 2)"),
    "2025-09": ("2025/Taxes 2025/Cosmedici Laser/Accounted Statements/Septemember Master Sheet.xlsx", "Expenses Sheet"),
    "2025-10": ("2025/Taxes 2025/Cosmedici Laser/Accounted Statements/October Master Sheet (up to date).xlsb.xlsx", "Expenses Sheet"),
    "2025-11": ("2025/Taxes 2025/Cosmedici Laser/Accounted Statements/November Master Sheet (up to date).xlsb - Copy.xlsx", "Expenses Sheet"),
    "2025-12": ("2025/Taxes 2025/Cosmedici Laser/Accounted Statements/December Master Sheet..xlsx", "Expenses Sheet"),
    "2026-01": ("2026/01_2026/Jan Master Sheet..xlsx", "Expenses Sheet"),
    "2026-02": ("2026/02_2026/Feb Master Sheet.xlsx", "Expenses Sheet"),
    "2026-03": ("2026/03_2026/March Master Sheet.xlsx", "Expenses Sheet"),
    "2026-04": ("2026/04_2026/April Master Sheet.xlsx", "Expenses Sheet"),
    "2026-05": ("2026/05_2026/May Master Sheet.xlsx", "Expenses Sheet"),
    "2026-06": ("2026/06_2026/June Master Sheet.xlsx", "Expenses Sheet"),
    "2026-07": ("2026/07_2026/July Master Sheet.xlsx", "Expenses Sheet"),
}

REVENUE_SOURCES = {
    "2025-08": ("2025/Revenue 2025/Service_Revenue_08_25.xlsx", None),
    "2025-09": ("2025/Taxes 2025/Cosmedici Laser/Accounted Statements/Septemember Master Sheet.xlsx", "Service Revenue Sheet"),
    "2025-10": ("2025/Taxes 2025/Cosmedici Laser/Accounted Statements/October Master Sheet (up to date).xlsb.xlsx", "Service Revenue Sheet"),
    "2025-11": ("2025/Taxes 2025/Cosmedici Laser/Accounted Statements/November Master Sheet (up to date).xlsb - Copy.xlsx", "Service Revenue Sheet"),
    "2025-12": ("2025/Taxes 2025/Cosmedici Laser/Accounted Statements/December Master Sheet..xlsx", "Service Revenue Sheet"),
    "2026-01": ("2026/01_2026/Jan Master Sheet..xlsx", "Service Revenue Sheet"),
    "2026-02": ("2026/02_2026/Feb Master Sheet.xlsx", "Service Revenue Sheet"),
    "2026-03": ("2026/03_2026/March Master Sheet.xlsx", "Service Revenue Sheet"),
    "2026-04": ("2026/04_2026/April Master Sheet.xlsx", "Service Revenue Sheet"),
    "2026-05": ("2026/05_2026/May Master Sheet.xlsx", "Service Revenue Sheet"),
    "2026-06": ("2026/06_2026/June Master Sheet.xlsx", "Service Revenue Sheet"),
    "2026-07": ("2026/07_2026/July Master Sheet.xlsx", "Service Revenue Sheet"),
}


def fix_year_typo(date_val: dt.date, month_key: str) -> dt.date:
    """Correct a mistyped year while leaving a genuinely correct date alone.

    Each source file covers exactly one calendar month (month_key, e.g.
    "2025-12"), so any row's real month should always match it. When a
    row's parsed month matches month_key's month but the year doesn't,
    the year was mistyped (e.g. some rows typed with a stray year near a
    month boundary) - correct the year rather than hardcoding each bad
    value found so far, since this class of typo keeps turning up in new
    places.
    """
    expected_year, expected_month = (int(p) for p in month_key.split("-"))
    if date_val.month == expected_month and date_val.year != expected_year:
        return date_val.replace(year=expected_year)
    return date_val


def _is_artifact_row(date_val, vendor_or_desc_present: bool) -> bool:
    """A row with no date and no vendor/description is a summary/total
    artifact baked into the sheet, not a real transaction."""
    return date_val is None and not vendor_or_desc_present


def load_expenses_month(raw_dir: Path, month: str) -> pd.DataFrame:
    rel_path, sheet_name = EXPENSE_SOURCES[month]
    wb = openpyxl.load_workbook(raw_dir / rel_path, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for r in range(2, ws.max_row + 1):
        date_val, week_val, vendor, category, description, amount, payment_method = (
            ws.cell(row=r, column=c).value for c in range(1, 8)
        )
        if all(v is None for v in (date_val, week_val, vendor, category, description, amount, payment_method)):
            continue

        has_vendor_or_desc = bool(vendor) or bool(description)
        date_estimated = False

        if date_val is None:
            # A handful of real rows are missing a date but have a known
            # estimated date on file (see config/private_cleaning_overrides.py).
            # These 3-ish rows can have the vendor name sitting in the
            # description field instead of vendor - try both fields.
            vendor_key = (month, str(vendor).strip() if vendor else None, amount)
            desc_key = (month, str(description).strip() if description else None, amount)

            if vendor_key in ESTIMATED_DATES:
                date_val = ESTIMATED_DATES[vendor_key]
                date_estimated = True
            elif desc_key in ESTIMATED_DATES:
                date_val = ESTIMATED_DATES[desc_key]
                date_estimated = True
            elif _is_artifact_row(date_val, has_vendor_or_desc):
                continue  # drop: summary/total artifact row
            else:
                # Unexpected: a row with no date, no known estimate, but has
                # vendor/description info that doesn't match anything on
                # file. Don't silently drop - raise so it gets investigated
                # instead of quietly disappearing. If you're seeing this on
                # real data, add the row to config/private_cleaning_overrides.py.
                raise ValueError(
                    f"Unmatched no-date row in {month}: vendor={vendor!r}, "
                    f"description={description!r}, amount={amount!r}. "
                    f"Add it to ESTIMATED_DATES in config/private_cleaning_overrides.py "
                    f"or confirm it's an artifact."
                )

        if isinstance(date_val, dt.datetime):
            date_val = date_val.date()
        if isinstance(date_val, dt.date):
            date_val = fix_year_typo(date_val, month)

        clean_category = CATEGORY_CROSSWALK.get(
            str(category).strip() if category else None, "Other/Non-Operating"
        )

        rows.append(
            {
                "date": date_val,
                "date_estimated": date_estimated,
                "vendor_name": vendor,
                "category": clean_category,
                "category_raw": category,
                "description": description,
                "amount": amount,
                "payment_method": payment_method,
                "source_month": month,
            }
        )

    return pd.DataFrame(rows)


def load_revenue_month(raw_dir: Path, month: str) -> pd.DataFrame:
    rel_path, sheet_name = REVENUE_SOURCES[month]
    wb = openpyxl.load_workbook(raw_dir / rel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows = []
    for r in range(2, ws.max_row + 1):
        date_val, revenue_source, amount, payment_method = (
            ws.cell(row=r, column=c).value for c in range(1, 5)
        )
        if all(v is None for v in (date_val, revenue_source, amount, payment_method)):
            continue
        if date_val is None:
            # No known estimated-date cases on the revenue side (all were
            # confirmed artifacts during the Sprint 2 audit) - drop.
            continue

        if isinstance(date_val, dt.datetime):
            date_val = date_val.date()
        if isinstance(date_val, dt.date):
            date_val = fix_year_typo(date_val, month)

        rows.append(
            {
                "date": date_val,
                "revenue_source": revenue_source,
                "amount": amount,
                "payment_method": payment_method,
                "source_month": month,
            }
        )

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# revenue_source label cleanup - typos/whitespace variants found in the
# Sprint 2 validation pass. These are payment-platform product names
# (Square, Fresha), not private business information, so they're safe to
# keep in the public script.
# ---------------------------------------------------------------------------

REVENUE_SOURCE_FIXES = {
    "Sqaure": "Square",
    "square": "Square",
    "SQUARE": "Square",
    "fresha": "Fresha",
    "FRESHA": "Fresha",
}


def normalize_revenue_source(value):
    if value is None:
        return value
    text = str(value).strip()
    return REVENUE_SOURCE_FIXES.get(text, text)


def apply_vendor_category_fixes(row):
    """Apply real-data vendor-based category corrections loaded from
    config/private_cleaning_overrides.py (empty/no-op if that file isn't
    present - e.g. when running against synthetic data)."""
    if not VENDOR_CATEGORY_OVERRIDES:
        return row["category"]

    vendor = str(row["vendor_name"]) if pd.notna(row["vendor_name"]) else ""
    desc = str(row["description"]) if pd.notna(row["description"]) else ""

    for rule in VENDOR_CATEGORY_OVERRIDES:
        if not re.search(rule["vendor_pattern"], vendor, re.I):
            continue
        desc_includes = rule.get("desc_includes")
        if desc_includes and desc_includes.lower() not in desc.lower():
            continue
        desc_excludes = rule.get("desc_excludes")
        if desc_excludes and desc_excludes.lower() in desc.lower():
            continue
        return rule["target_category"]

    return row["category"]


# Sprint 3 sparse-category rule: these categories have too few real rows
# (~10 or fewer) to reliably train a per-category prediction on. Per
# project decision, they're folded into Other/Non-Operating for MODELING
# purposes only - `category` keeps the true label for real
# bookkeeping/reporting, `category_for_model` is what the classifier is
# trained and evaluated on. These are business category names, not
# vendor/person names, so safe to keep here.
SPARSE_CATEGORIES_MERGE_TO_OTHER = {
    "Professional Services",
    "Licensing & Compliance",
}


def add_model_category_column(df: pd.DataFrame) -> pd.DataFrame:
    df["category_for_model"] = df["category"].apply(
        lambda c: "Other/Non-Operating" if c in SPARSE_CATEGORIES_MERGE_TO_OTHER else c
    )
    return df


def clean_all(raw_dir: Path):
    expense_frames = [load_expenses_month(raw_dir, m) for m in EXPENSE_SOURCES]
    revenue_frames = [load_revenue_month(raw_dir, m) for m in REVENUE_SOURCES]

    expenses = pd.concat(expense_frames, ignore_index=True)
    revenue = pd.concat(revenue_frames, ignore_index=True)

    revenue["revenue_source"] = revenue["revenue_source"].apply(normalize_revenue_source)

    expenses["category"] = expenses.apply(apply_vendor_category_fixes, axis=1)
    expenses = add_model_category_column(expenses)

    expenses = expenses.sort_values("date").reset_index(drop=True)
    revenue = revenue.sort_values("date").reset_index(drop=True)

    expenses.insert(0, "expense_id", range(1, len(expenses) + 1))
    revenue.insert(0, "revenue_id", range(1, len(revenue) + 1))

    bad_categories = set(expenses["category"]) - set(FINAL_CATEGORIES)
    if bad_categories:
        raise ValueError(
            f"Categories found that aren't in FINAL_CATEGORIES: {bad_categories}. "
            f"Check CATEGORY_CROSSWALK for a missing legacy label."
        )

    min_date = dt.date(2025, 8, 1)
    max_date = dt.date(2026, 7, 31)
    for label, df in (("expenses", expenses), ("revenue", revenue)):
        out_of_range = df[(df["date"] < min_date) | (df["date"] > max_date)]
        if not out_of_range.empty:
            raise ValueError(
                f"{len(out_of_range)} {label} row(s) fall outside the expected "
                f"{min_date}..{max_date} range - likely an uncaught year typo:\n"
                f"{out_of_range[['date', 'source_month']].to_string()}"
            )

    return expenses, revenue


def print_summary(expenses, revenue):
    print(f"Expenses: {len(expenses)} rows, {expenses['date'].min()} to {expenses['date'].max()}")
    print(f"  Estimated dates: {int(expenses['date_estimated'].sum())} row(s)")
    print("  By category:")
    for cat, total in expenses.groupby("category")["amount"].sum().sort_values(ascending=False).items():
        print(f"    {cat}: ${total:,.2f}")
    print("  By category_for_model:")
    for cat, count in expenses["category_for_model"].value_counts().items():
        print(f"    {cat}: {count} rows")

    print(f"\nRevenue: {len(revenue)} rows, {revenue['date'].min()} to {revenue['date'].max()}")
    print("  By source:")
    for src, total in revenue.groupby("revenue_source")["amount"].sum().sort_values(ascending=False).items():
        print(f"    {src}: ${total:,.2f}")


def main():
    parser = argparse.ArgumentParser(description="Clean Cosmedici expense and revenue data.")
    parser.add_argument(
        "--raw-dir",
        type=Path,
        required=True,
        help="Folder that directly contains the '2025' and '2026' export subfolders.",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path("data/private"),
        help="Where to write cleaned_expenses.csv / cleaned_revenue.csv (default: data/private).",
    )
    args = parser.parse_args()

    expenses, revenue = clean_all(args.raw_dir)
    print_summary(expenses, revenue)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    expenses.to_csv(args.out_dir / "cleaned_expenses.csv", index=False)
    revenue.to_csv(args.out_dir / "cleaned_revenue.csv", index=False)
    print(f"\nWrote {args.out_dir / 'cleaned_expenses.csv'}")
    print(f"Wrote {args.out_dir / 'cleaned_revenue.csv'}")


if __name__ == "__main__":
    main()
